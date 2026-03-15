"""
generate_glossary_template.py
Generates Business_Glossary_Import_Template.xlsx for the team to fill
from the old glossary tool, ready for import into SC_QAWS.

Run:  python generate_glossary_template.py
Requires: pip install openpyxl
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Colour palette ────────────────────────────────────────────────────────────
RED_FILL    = PatternFill("solid", fgColor="C00000")   # required header
AMBER_FILL  = PatternFill("solid", fgColor="ED7D31")   # optional header
BLUE_FILL   = PatternFill("solid", fgColor="1F4E79")   # section header
GREY_FILL   = PatternFill("solid", fgColor="D9D9D9")   # read-only / notes
GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")   # data rows (even)
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")   # data rows (odd)
INSTR_FILL  = PatternFill("solid", fgColor="FFF2CC")   # instruction rows

WHITE_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BLACK_FONT  = Font(name="Calibri", bold=True, color="000000", size=11)
BODY_FONT   = Font(name="Calibri", size=10)
INSTR_FONT  = Font(name="Calibri", size=10, italic=True, color="7F7F7F")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Column definitions ────────────────────────────────────────────────────────
# (header, width, required, note, maps_to)
COLUMNS = [
    # ── Hierarchy ──────────────────────────────────────────────────────────
    ("Glossary Name (EN)",       28, True,
     "Top-level glossary name (e.g. 'Statistical Business Register')",
     "glossary.primaryname (level 0)"),

    ("Glossary Name (AR)",       28, False,
     "Arabic name of the glossary",
     "custom_field 120 @ level 0"),

    ("Topic Name (EN)",          28, True,
     "Level-1 topic / domain (e.g. 'Economy')",
     "glossary.primaryname (level 1)"),

    ("Topic Name (AR)",          28, False,
     "Arabic topic name",
     "custom_field 120 @ level 1"),

    ("Theme Name (EN)",          28, True,
     "Level-2 theme / sub-topic (e.g. 'Trade')",
     "glossary.primaryname (level 2)"),

    ("Theme Name (AR)",          28, False,
     "Arabic theme name",
     "custom_field 120 @ level 2"),

    # ── Dataset (optional level between theme and term) ───────────────────
    ("Dataset Name (EN)",        28, False,
     "Leave blank if term sits directly under a theme. Fill only when an "
     "intermediate dataset node exists (type=9).",
     "glossary.primaryname (level 3, type=9)"),

    ("Dataset Name (AR)",        28, False,
     "Arabic dataset name (if applicable)",
     "custom_field 120 @ level 3 type=9"),

    # ── Term ──────────────────────────────────────────────────────────────
    ("Term Name (EN)",           30, True,
     "English term name",
     "glossary.primaryname (term node)"),

    ("Term Name (AR)",           30, False,
     "Arabic term name",
     "custom_field 120 @ term node"),

    ("Term Definition (EN)",     50, False,
     "Full English definition",
     "glossary.description (term node)"),

    ("Term Definition (AR)",     50, False,
     "Full Arabic definition",
     "custom_field 121 @ term node"),

    ("Term Source",              30, False,
     "Reference / source of the definition",
     "custom_field 146 @ term node"),

    # ── Classification & visibility ───────────────────────────────────────
    ("Axon Viewing",             18, True,
     "Must be exactly: Public  OR  Private",
     "glossary.ispublic  (1=Public, 0=Private)"),

    ("Term Status",              20, False,
     "Active / Pending Review / Retired  (leave blank → Pending Review)",
     "glossary.status  (1=Active, 3=Pending Review)"),

    ("Security Classification",  25, False,
     "e.g. Public, Restricted, Confidential — must match lookup table value",
     "glossary.securityclassification (FK to security_classification)"),

    # ── Ref numbers (for import script – generate if blank) ──────────────
    ("Term Ref",                 20, False,
     "Unique reference code from old tool.  Leave blank to auto-generate.",
     "glossary.refnumber (term node)"),

    ("Parent Term Ref",          20, False,
     "refnumber of parent node (theme or dataset). "
     "Leave blank if you filled Dataset/Theme columns — import script will resolve.",
     "glossary.parent_id resolved via refnumber"),

    # ── Notes ─────────────────────────────────────────────────────────────
    ("Notes / Remarks",          35, False,
     "Any migration notes — NOT imported to DB",
     "— (ignored by import script)"),
]


def style_header(cell, required):
    cell.fill   = RED_FILL if required else AMBER_FILL
    cell.font   = WHITE_FONT
    cell.border = BORDER
    cell.alignment = CENTER


def add_data_sheet(wb):
    ws = wb.active
    ws.title = "Glossary Data"

    # ── Row 1: section labels ─────────────────────────────────────────────
    sections = [
        (1,  2,  "HIERARCHY"),
        (3,  4,  "TOPIC"),
        (5,  6,  "THEME"),
        (7,  8,  "DATASET (optional)"),
        (9,  13, "TERM"),
        (14, 16, "CLASSIFICATION"),
        (17, 18, "REF NUMBERS"),
        (19, 19, "NOTES"),
    ]
    for start_col, end_col, label in sections:
        ws.merge_cells(start_row=1, start_column=start_col,
                       end_row=1,   end_column=end_col)
        cell = ws.cell(row=1, column=start_col, value=label)
        cell.fill      = BLUE_FILL
        cell.font      = WHITE_FONT
        cell.alignment = CENTER
        cell.border    = BORDER

    # ── Row 2: column headers ─────────────────────────────────────────────
    for col_i, (header, width, required, note, maps_to) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_i, value=header)
        style_header(cell, required)
        ws.column_dimensions[get_column_letter(col_i)].width = width
        # store the mapping note as a comment-like value in row 3
        note_cell = ws.cell(row=3, column=col_i, value=f"↳ {maps_to}")
        note_cell.fill      = GREY_FILL
        note_cell.font      = INSTR_FONT
        note_cell.alignment = LEFT
        note_cell.border    = BORDER

    # ── Row 3 height ─────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 28

    # ── Data rows 4-203 (200 rows for the team) ───────────────────────────
    for row_i in range(4, 204):
        fill = GREEN_FILL if row_i % 2 == 0 else WHITE_FILL
        for col_i in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_i, column=col_i)
            cell.fill      = fill
            cell.font      = BODY_FONT
            cell.alignment = LEFT
            cell.border    = BORDER

    # ── Data Validations ─────────────────────────────────────────────────
    # Axon Viewing  → col 14
    dv_viewing = DataValidation(
        type="list",
        formula1='"Public,Private"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error='Must be "Public" or "Private"',
        showDropDown=False,
    )
    dv_viewing.sqref = f"N4:N203"
    ws.add_data_validation(dv_viewing)

    # Term Status → col 15
    dv_status = DataValidation(
        type="list",
        formula1='"Active,Pending Review,Retired"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error='Choose: Active, Pending Review, or Retired',
        showDropDown=False,
    )
    dv_status.sqref = "O4:O203"
    ws.add_data_validation(dv_status)

    # ── Freeze panes below header ─────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Auto-filter on row 2 ──────────────────────────────────────────────
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}203"

    return ws


def add_instructions_sheet(wb):
    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80

    rows = [
        ("PURPOSE", "Fill one row per Term from the old Axon/Glossary tool."),
        ("", "The import script (load_glossary.py) reads this file and inserts"),
        ("", "records into SC_QAWS.GLOSSARY and SC_QAWS.CUSTOM_FIELD."),
        ("", ""),
        ("REQUIRED COLUMNS", "Red headers = must be filled. Amber = optional but recommended."),
        ("", ""),
        ("HIERARCHY", "Each row must have: Glossary Name → Topic → Theme → Term Name."),
        ("", "If the old tool had a Dataset level between Theme and Term, fill those columns too."),
        ("", ""),
        ("AXON VIEWING", 'Use exactly "Public" or "Private" (dropdown enforced).'),
        ("", "Public → visible in the glossary portal."),
        ("", "Private → hidden (ispublic=0)."),
        ("", ""),
        ("TERM STATUS", '"Active" → published. "Pending Review" → draft (default if left blank).'),
        ("", ""),
        ("REF NUMBERS", "Copy Term Ref from the old tool if available."),
        ("", "If blank, the import script auto-generates a sequential ref."),
        ("", "Parent Term Ref: only needed when import cannot resolve from hierarchy columns."),
        ("", ""),
        ("ARABIC FIELDS", "Paste Arabic text directly. The sheet supports RTL characters."),
        ("", ""),
        ("AFTER FILLING",
         "1. Save as .xlsx (keep as-is)."),
        ("", "2. Hand the file back to the Data Steward / DBA."),
        ("", "3. DBA runs:  python load_glossary.py Business_Glossary_Import_Template.xlsx"),
        ("", "4. Refresh the MV:  EXEC DBMS_MVIEW.REFRESH('SC_QAWS.BUSINESS_GLOSSARY','C');"),
        ("", ""),
        ("QUESTIONS", "Contact the QAWS team or raise a ticket in the usual channel."),
    ]

    ws.cell(row=1, column=1, value="Business Glossary — Migration Fill Guide").font = Font(
        name="Calibri", bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 24

    for i, (label, text) in enumerate(rows, start=2):
        lc = ws.cell(row=i, column=1, value=label)
        tc = ws.cell(row=i, column=2, value=text)
        if label:
            lc.font      = Font(name="Calibri", bold=True, size=10, color="1F4E79")
            lc.fill      = INSTR_FILL
            lc.alignment = LEFT
        else:
            lc.font      = BODY_FONT
        tc.font      = BODY_FONT
        tc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[i].height = 16

    return ws


def add_valid_values_sheet(wb):
    ws = wb.create_sheet("Valid Values")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    data = [
        ("Field",                  "Allowed Values"),
        ("Axon Viewing",           "Public"),
        ("",                       "Private"),
        ("Term Status",            "Active"),
        ("",                       "Pending Review"),
        ("",                       "Retired"),
        ("Security Classification","Public"),
        ("",                       "Restricted"),
        ("",                       "Confidential"),
        ("",                       "Top Secret"),
        ("",                       "(match exactly with SC_QAWS.SECURITY_CLASSIFICATION table)"),
    ]

    for i, (col_a, col_b) in enumerate(data, start=1):
        ca = ws.cell(row=i, column=1, value=col_a)
        cb = ws.cell(row=i, column=2, value=col_b)
        if i == 1:
            ca.fill = BLUE_FILL; ca.font = WHITE_FONT
            cb.fill = BLUE_FILL; cb.font = WHITE_FONT
        else:
            ca.font = BODY_FONT
            cb.font = BODY_FONT
        ca.border = BORDER; cb.border = BORDER
        ca.alignment = LEFT;  cb.alignment = LEFT

    return ws


def main():
    wb = openpyxl.Workbook()
    add_data_sheet(wb)
    add_instructions_sheet(wb)
    add_valid_values_sheet(wb)

    out = "Business_Glossary_Import_Template.xlsx"
    wb.save(out)
    print(f"Done. Created: {out}")
    print(f"   Sheets : Glossary Data | Instructions | Valid Values")
    print(f"   Rows   : 200 data rows (rows 4-203), headers in rows 1-3")
    print(f"   Columns: {len(COLUMNS)}")


if __name__ == "__main__":
    main()
